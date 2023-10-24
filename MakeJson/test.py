import json
from openpyxl import Workbook

with open('Festival.json', 'r', encoding='utf-8') as f :
    dataList = json.load(f)

print(len(dataList))
wb = Workbook()
ws = wb.active

row = 0
  
for i in range(len(dataList)):
    ob = dataList[i]
    if i == 1 :
        ws[f"B{i}"] = 'RECOMMEND_TRAVEL_ADDRESS'
        ws[f"C{i}"] = 'RECOMMEND_TRAVEL_EVENTSTARTDATE'
        ws[f"D{i}"] = 'RECOMMEND_TRAVEL_EVENTENDDATE'
        ws[f"E{i}"] = 'RECOMMEND_TRAVEL_IMG'
        ws[f"G{i}"] = 'RECOMMEND_TRAVEL_MAPX'
        ws[f"H{i}"] = 'RECOMMEND_TRAVEL_MAPY'
        ws[f"I{i}"] = 'RECOMMEND_TRAVEL_MAPLEVEL'
        ws[f"J{i}"] = 'RECOMMEND_TRAVEL_TEL'
        ws[f"K{i}"] = 'RECOMMEND_TRAVEL_TITLE'
        ws[f"F{i}"] = 'RECOMMEND_TRAVEL_CONTENT_ID'
    
    ws[f"B{i + 1}"] = ob['addr1'] + ob['addr2']
    ws[f"C{i + 1}"] = ob['eventstartdate']
    ws[f"D{i + 1}"] = ob['eventenddate']
    ws[f"E{i + 1}"] = ob['firstimage']
    ws[f"G{i + 1}"] = ob['mapx']
    ws[f"H{i + 1}"] = ob['mapy']
    ws[f"I{i + 1}"] = ob['mlevel']
    ws[f"J{i + 1}"] = ob['tel']
    ws[f"K{i + 1}"] = ob['title']
    ws[f"F{i + 1}"] = ob['contentid']
    

wb.save('Festival.xlsx')